"""
Servicio para generar reportes en Excel
"""
from io import BytesIO
from datetime import date
from typing import List
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelService:
    """Genera reportes en formato Excel"""

    @staticmethod
    def _aplicar_estilos_encabezado(ws, num_columnas: int, fila: int = 1):
        """Aplica estilos a la fila de encabezados"""
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, num_columnas + 1):
            cell = ws.cell(row=fila, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    @staticmethod
    def _ajustar_anchos(ws):
        """Ajusta el ancho de columnas automaticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @classmethod
    def generar_reporte_empleados(cls, empleados, empresa_nombre: str) -> BytesIO:
        """Genera reporte de empleados en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Empleados"

        # Titulo
        ws.append([f'Reporte de Empleados - {empresa_nombre}'])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.append([f'Generado: {date.today().strftime("%d/%m/%Y")}'])
        ws.append([])

        # Encabezados
        headers = [
            'Nombre Completo', 'RFC', 'CURP', 'Puesto', 'Departamento',
            'Fecha Ingreso', 'Antiguedad', 'Salario Diario', 'Estado', 'Jefe Directo'
        ]
        ws.append(headers)
        cls._aplicar_estilos_encabezado(ws, len(headers), fila=4)

        # Datos
        for emp in empleados:
            antiguedad = ''
            if hasattr(emp, 'antiguedad'):
                antiguedad = emp.antiguedad
            elif emp.fecha_ingreso:
                dias = (date.today() - emp.fecha_ingreso).days
                anos = dias // 365
                meses = (dias % 365) // 30
                antiguedad = f"{anos}a {meses}m"

            ws.append([
                emp.nombre_completo,
                emp.rfc or '',
                emp.curp or '',
                emp.puesto or '',
                emp.departamento or '',
                emp.fecha_ingreso.strftime('%d/%m/%Y') if emp.fecha_ingreso else '',
                antiguedad,
                float(emp.salario_diario) if emp.salario_diario else 0,
                emp.estado,
                emp.jefe_directo.nombre_completo if emp.jefe_directo else '',
            ])

        cls._ajustar_anchos(ws)

        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    @classmethod
    def generar_reporte_nomina(cls, periodo, recibos) -> BytesIO:
        """Genera reporte de nomina en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Nomina {periodo.numero_periodo}"

        # Informacion del periodo
        ws.append([f'Reporte de Nomina'])
        ws['A1'].font = Font(bold=True, size=14)

        ws.append([f'Periodo: {periodo.tipo_periodo} {periodo.numero_periodo}/{periodo.ano}'])
        ws.append([f'Fecha: {periodo.fecha_inicio.strftime("%d/%m/%Y")} - {periodo.fecha_fin.strftime("%d/%m/%Y")}'])
        ws.append([f'Estado: {periodo.estado}'])
        ws.append([])

        # Encabezados
        headers = [
            'Empleado', 'RFC', 'Dias Trabajados', 'Salario Diario',
            'Percepciones', 'ISR', 'IMSS', 'Otras Deducciones', 'Total Deducciones', 'Neto'
        ]
        ws.append(headers)
        fila_headers = 6
        cls._aplicar_estilos_encabezado(ws, len(headers), fila=fila_headers)

        # Datos
        fila_inicial = fila_headers + 1
        for i, recibo in enumerate(recibos):
            otras_deducciones = float(recibo.total_deducciones or 0) - float(recibo.isr_retenido or 0) - float(recibo.cuota_imss_obrera or 0)
            ws.append([
                recibo.empleado.nombre_completo if recibo.empleado else 'N/A',
                recibo.empleado.rfc if recibo.empleado else '',
                recibo.dias_trabajados or 0,
                float(recibo.salario_diario or 0),
                float(recibo.total_percepciones or 0),
                float(recibo.isr_retenido or 0),
                float(recibo.cuota_imss_obrera or 0),
                otras_deducciones if otras_deducciones > 0 else 0,
                float(recibo.total_deducciones or 0),
                float(recibo.neto_a_pagar or 0),
            ])

        # Totales
        fila_totales = fila_inicial + len(list(recibos))
        ws.append([])
        ws.append([
            'TOTALES', '', '', '',
            f'=SUM(E{fila_inicial}:E{fila_totales-1})',
            f'=SUM(F{fila_inicial}:F{fila_totales-1})',
            f'=SUM(G{fila_inicial}:G{fila_totales-1})',
            f'=SUM(H{fila_inicial}:H{fila_totales-1})',
            f'=SUM(I{fila_inicial}:I{fila_totales-1})',
            f'=SUM(J{fila_inicial}:J{fila_totales-1})',
        ])

        # Estilo para totales
        for col in range(1, 11):
            cell = ws.cell(row=fila_totales + 1, column=col)
            cell.font = Font(bold=True)

        cls._ajustar_anchos(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    @classmethod
    def generar_reporte_incidencias(cls, incidencias, mes: int, ano: int) -> BytesIO:
        """Genera reporte de incidencias en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Incidencias {mes}-{ano}"

        # Titulo
        ws.append([f'Reporte de Incidencias - {mes}/{ano}'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([f'Generado: {date.today().strftime("%d/%m/%Y")}'])
        ws.append([])

        headers = [
            'Empleado', 'Tipo', 'Fecha Inicio', 'Fecha Fin',
            'Cantidad', 'Descripcion', 'Aplicado'
        ]
        ws.append(headers)
        cls._aplicar_estilos_encabezado(ws, len(headers), fila=4)

        for inc in incidencias:
            ws.append([
                inc.empleado.nombre_completo if inc.empleado else 'N/A',
                inc.tipo or '',
                inc.fecha_inicio.strftime('%d/%m/%Y') if inc.fecha_inicio else '',
                inc.fecha_fin.strftime('%d/%m/%Y') if inc.fecha_fin else '',
                float(inc.cantidad) if inc.cantidad else 1,
                inc.descripcion or '',
                'Si' if getattr(inc, 'aplicado', False) else 'No',
            ])

        cls._ajustar_anchos(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    @classmethod
    def generar_reporte_solicitudes(cls, solicitudes, empresa_nombre: str) -> BytesIO:
        """Genera reporte de solicitudes"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes"

        # Titulo
        ws.append([f'Reporte de Solicitudes - {empresa_nombre}'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([f'Generado: {date.today().strftime("%d/%m/%Y")}'])
        ws.append([])

        headers = [
            'Fecha', 'Solicitante', 'Tipo', 'Titulo',
            'Estado', 'Prioridad', 'Aprobador', 'Fecha Respuesta'
        ]
        ws.append(headers)
        cls._aplicar_estilos_encabezado(ws, len(headers), fila=4)

        for sol in solicitudes:
            ws.append([
                sol.created_at.strftime('%d/%m/%Y') if sol.created_at else '',
                sol.solicitante.nombre_completo if sol.solicitante else '',
                sol.get_tipo_display() if hasattr(sol, 'get_tipo_display') else sol.tipo,
                sol.titulo or '',
                sol.get_estado_display() if hasattr(sol, 'get_estado_display') else sol.estado,
                sol.prioridad or 'media',
                sol.aprobador.email if sol.aprobador else '',
                sol.updated_at.strftime('%d/%m/%Y') if sol.estado not in ['pendiente', 'borrador'] and sol.updated_at else '',
            ])

        cls._ajustar_anchos(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    @classmethod
    def generar_reporte_auditoria(cls, registros) -> BytesIO:
        """Genera reporte de auditoria"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria"

        # Titulo
        ws.append(['Reporte de Auditoria'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([f'Generado: {date.today().strftime("%d/%m/%Y")}'])
        ws.append([])

        headers = [
            'Fecha/Hora', 'Usuario', 'Accion', 'Descripcion',
            'Modelo', 'IP'
        ]
        ws.append(headers)
        cls._aplicar_estilos_encabezado(ws, len(headers), fila=4)

        for r in registros:
            ws.append([
                r.created_at.strftime('%d/%m/%Y %H:%M') if r.created_at else '',
                r.usuario.email if r.usuario else 'Sistema',
                r.get_accion_display() if hasattr(r, 'get_accion_display') else r.accion,
                r.descripcion[:100] if r.descripcion else '',
                r.modelo or '',
                r.ip_address or '',
            ])

        cls._ajustar_anchos(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
